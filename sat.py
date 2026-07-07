import sys
import unicodedata
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def normalize(s):
    if s is None:
        return ''
    s = str(s).replace('\n', ' ')
    s = unicodedata.normalize('NFD', s)
    s = ''.join((ch for ch in s if unicodedata.category(ch) != 'Mn'))
    s = s.replace('Đ', 'D').replace('đ', 'd')
    return ' '.join(s.lower().split())

def find_students(ws, name_row):
    cols, col, empty = ([], 3, 0)
    while empty < 3:
        v = ws.cell(row=name_row, column=col).value
        if v is not None and str(v).strip():
            cols.append(col)
            empty = 0
        else:
            empty += 1
        col += 1
    return cols

def analyze(path):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    name_row = exam_row = None
    skill_rows = {}
    section = ''
    for r in range(1, (ws.max_row or 200) + 1):
        raw = ws.cell(row=r, column=2).value
        if raw is None:
            continue
        n = normalize(raw)
        if str(raw).strip().startswith('▶'):
            if 'ket qua bai thi — sat' in n:
                section = 'SAT'
            elif 'ket qua bai thi — ielts' in n:
                section = 'IELTS'
            else:
                section = ''
            continue
        if n.startswith('ho va ten'):
            name_row = r
        elif n.startswith('ky thi'):
            exam_row = r
        elif section and ('— dung' in n or '— tong so cau' in n):
            label = str(raw).replace('\n', ' ').strip()
            skill = label.rsplit('—', 1)[0].strip()
            kind = 'correct' if '— dung' in n else 'total'
            skill_rows.setdefault((section, skill), {})[kind] = r
    if name_row is None:
        sys.exit("Error: could not find the 'Họ và tên' row — wrong template?")
    for col in find_students(ws, name_row):
        name = str(ws.cell(row=name_row, column=col).value).strip()
        exam = str(ws.cell(row=exam_row, column=col).value or '').strip().upper()
        if exam not in ('SAT', 'IELTS'):
            print(f'\n[{get_column_letter(col)}] {name}: exam type {exam!r} is not SAT/IELTS — skipped.')
            continue
        results = []
        for (sec, skill), rows in skill_rows.items():
            if sec != exam:
                continue
            correct = ws.cell(row=rows.get('correct', 0), column=col).value if rows.get('correct') else None
            total = ws.cell(row=rows.get('total', 0), column=col).value if rows.get('total') else None
            try:
                correct = int(correct) if correct is not None else None
                total = int(total) if total is not None else None
            except (TypeError, ValueError):
                continue
            if total is None or total <= 0:
                continue
            if correct is None:
                correct = 0
            if correct > total:
                print(f'  ! {name} / {skill}: correct {correct} > total {total} — skipped, check the data.')
                continue
            results.append((skill, correct, total, correct / total * 100))
        print(f'\n=== {name} ({exam}) ===')
        if not results:
            print('  No exam data entered for this student.')
            continue
        results.sort(key=lambda x: x[3])
        for skill, correct, total, pct in results:
            bar = '█' * round(pct / 5)
            print(f'  {pct:5.1f}%  {bar:<20s} {skill}  ({correct}/{total})')
        overall = sum((r[1] for r in results)) / sum((r[2] for r in results)) * 100
        print(f'  Overall accuracy: {overall:.1f}%  |  Weakest first — start revising from the top.')
if __name__ == '__main__':
    analyze(sys.argv[1] if len(sys.argv) > 1 else 'student_intake.xlsx')
